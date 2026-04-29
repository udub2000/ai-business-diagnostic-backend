from __future__ import annotations

import csv
import json
import os
import re
import random
import time
from pathlib import Path
from typing import Any, Dict, Generator, Optional
from uuid import uuid4

import anthropic
from dotenv import load_dotenv
from fastapi import Depends, FastAPI, File, Form, HTTPException, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.db.database import Base, SessionLocal, engine
from app.db.repository import create_assessment, get_assessment_by_id, list_recent_assessments

load_dotenv()
Base.metadata.create_all(bind=engine)

app = FastAPI(title="AI Business Diagnostic API", version="v2")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://127.0.0.1:3000", "https://theuniikyetruth.com"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def get_db() -> Generator[Session, None, None]:
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


class AssessmentInput(BaseModel):
    business_type: str
    location: str
    owner_name: Optional[str] = None
    seasonality_context: Optional[str] = None
    analysis_mode: str = "auto"
    mode: str = "guided"
    leads_per_week: Optional[float] = None
    leads_estimate: Optional[str] = None
    schedule_fullness: Optional[str] = None
    backlog_days: Optional[float] = None
    backlog_bucket: Optional[str] = None
    conversion_rate: Optional[float] = None
    quotes_to_jobs_bucket: Optional[str] = None
    sales_behavior: Optional[str] = None
    response_speed: Optional[str] = None
    crews: Optional[int] = None
    jobs_per_crew_per_day: Optional[float] = None
    busy_day_description: Optional[str] = None
    turned_down_work: Optional[str] = None
    utilization_rate: Optional[float] = None
    can_take_work_tomorrow: Optional[str] = None
    average_job_time_hours: Optional[float] = None
    travel_time_minutes: Optional[float] = None
    scheduling_efficiency: Optional[str] = None
    time_loss_source: Optional[str] = None
    admin_drag: Optional[str] = None
    monthly_revenue: Optional[float] = None
    monthly_revenue_range: Optional[str] = None
    labor_cost_pct: Optional[float] = None
    profitability_feeling: Optional[str] = None
    gross_margin_pct: Optional[float] = None
    pricing_adequacy: Optional[str] = None
    cash_stress: Optional[str] = None
    income_statement_file_name: Optional[str] = None
    income_statement_file_path: Optional[str] = None
    balance_sheet_file_name: Optional[str] = None
    balance_sheet_file_path: Optional[str] = None
    cash_flow_statement_file_name: Optional[str] = None
    cash_flow_statement_file_path: Optional[str] = None
    financial_uploads_present: bool = False
    financial_metrics: Optional[Dict[str, Any]] = None
    financial_parse_summary: Optional[str] = None
    financial_parse_debug: Optional[Dict[str, Any]] = None
    notes: Optional[str] = None


def get_structured_input_completeness(payload: AssessmentInput) -> float:
    structured_fields = [
        payload.leads_per_week, payload.leads_estimate, payload.schedule_fullness,
        payload.backlog_days, payload.backlog_bucket,
        payload.conversion_rate, payload.quotes_to_jobs_bucket, payload.sales_behavior,
        payload.response_speed,
        payload.crews, payload.jobs_per_crew_per_day, payload.busy_day_description,
        payload.turned_down_work, payload.utilization_rate, payload.can_take_work_tomorrow,
        payload.average_job_time_hours, payload.travel_time_minutes, payload.scheduling_efficiency,
        payload.time_loss_source, payload.admin_drag,
        payload.monthly_revenue, payload.monthly_revenue_range, payload.labor_cost_pct,
        payload.profitability_feeling, payload.gross_margin_pct, payload.pricing_adequacy,
        payload.cash_stress,
    ]
    filled = sum(1 for x in structured_fields if x not in (None, "", 0))
    return filled / len(structured_fields)


def resolve_analysis_mode(payload: AssessmentInput) -> str:
    requested_mode = (payload.analysis_mode or "auto").strip().lower()
    notes = (payload.notes or "").strip()
    notes_len = len(notes)
    completeness = get_structured_input_completeness(payload)

    if requested_mode in {"notes_led", "hybrid", "structured"}:
        return requested_mode

    if notes_len >= 250 and completeness < 0.35:
        return "notes_led"
    if completeness >= 0.60:
        return "structured"
    return "hybrid"


def is_notes_dominant_mode(payload: AssessmentInput) -> bool:
    requested_mode = (payload.analysis_mode or "auto").strip().lower()
    notes = (payload.notes or "").strip()
    notes_len = len(notes)
    completeness = get_structured_input_completeness(payload)
    resolved_mode = resolve_analysis_mode(payload)

    if resolved_mode == "notes_led":
        return bool(notes)

    # Only auto mode is allowed to infer notes-led behavior.
    # Manual structured/hybrid overrides must be respected.
    if requested_mode == "auto":
        return notes_len >= 250 and completeness < 0.35

    return False


def should_ignore_blank_structured_fields(payload: AssessmentInput) -> bool:
    return resolve_analysis_mode(payload) == "notes_led"


def count_present(values: list[object]) -> int:
    return sum(1 for x in values if x not in (None, "", 0))


def build_score_statuses(payload: AssessmentInput) -> Dict[str, str]:
    if should_ignore_blank_structured_fields(payload) or is_notes_dominant_mode(payload):
        return {
            "demand": "insufficient_data",
            "conversion": "insufficient_data",
            "capacity": "insufficient_data",
            "operations": "insufficient_data",
            "financials": "insufficient_data",
        }

    category_inputs = {
        "demand": {
            "all": [
                payload.leads_per_week, payload.leads_estimate, payload.schedule_fullness,
                payload.backlog_days, payload.backlog_bucket,
            ],
            "direct": [payload.leads_per_week, payload.backlog_days],
        },
        "conversion": {
            "all": [
                payload.conversion_rate, payload.quotes_to_jobs_bucket,
                payload.sales_behavior, payload.response_speed,
            ],
            "direct": [payload.conversion_rate],
        },
        "capacity": {
            "all": [
                payload.crews, payload.jobs_per_crew_per_day, payload.busy_day_description,
                payload.turned_down_work, payload.utilization_rate,
                payload.can_take_work_tomorrow, payload.backlog_days, payload.backlog_bucket,
            ],
            "direct": [payload.crews, payload.jobs_per_crew_per_day, payload.utilization_rate],
        },
        "operations": {
            "all": [
                payload.jobs_per_crew_per_day, payload.travel_time_minutes,
                payload.scheduling_efficiency, payload.time_loss_source,
                payload.admin_drag, payload.average_job_time_hours,
            ],
            "direct": [payload.jobs_per_crew_per_day, payload.travel_time_minutes, payload.average_job_time_hours],
        },
        "financials": {
            "all": [
                payload.monthly_revenue, payload.monthly_revenue_range, payload.labor_cost_pct,
                payload.profitability_feeling, payload.gross_margin_pct,
                payload.pricing_adequacy, payload.cash_stress, payload.financial_metrics,
            ],
            "direct": [payload.monthly_revenue, payload.labor_cost_pct, payload.gross_margin_pct, payload.financial_metrics],
        },
    }

    statuses: Dict[str, str] = {}
    for category, parts in category_inputs.items():
        all_count = count_present(parts["all"])
        direct_count = count_present(parts["direct"])

        if all_count < 2:
            statuses[category] = "insufficient_data"
        elif direct_count >= 1:
            statuses[category] = "measured"
        else:
            statuses[category] = "estimated"

    return statuses


def sanitize_scores_for_display(scores: Dict[str, int], score_statuses: Dict[str, str]) -> Dict[str, Optional[int]]:
    return {
        category: (score if score_statuses.get(category) != "insufficient_data" else None)
        for category, score in scores.items()
    }


def build_score_summary_note(payload: AssessmentInput, score_statuses: Dict[str, str]) -> str:
    if should_ignore_blank_structured_fields(payload):
        return "This assessment is running in notes-led mode, so blank structured fields are being ignored and category scores are being treated as unavailable unless directly supported elsewhere."
    if is_notes_dominant_mode(payload):
        return "Limited structured input was provided, so category scores are being treated as unavailable while the report relies primarily on the notes field."

    insufficient = sum(1 for status in score_statuses.values() if status == "insufficient_data")
    estimated = sum(1 for status in score_statuses.values() if status == "estimated")

    if insufficient:
        return f"{insufficient} category score(s) are unavailable because there was not enough structured data to support them."
    if estimated:
        return f"{estimated} category score(s) are estimated from partial structured inputs rather than directly measured."
    return "Category scores are supported by enough structured input to be treated as reasonably grounded."


def score_from_bucket(value: Optional[str], mapping: Dict[str, int], default: int = 50) -> int:
    if not value:
        return default
    return mapping.get(value, default)


def clamp_score(value: float) -> int:
    return max(0, min(100, round(value)))


def score_leads(payload: AssessmentInput) -> int:
    if payload.leads_per_week is not None:
        x = payload.leads_per_week
        if x < 5:
            return 20
        if x < 15:
            return 40
        if x < 30:
            return 65
        if x < 50:
            return 82
        return 95
    return score_from_bucket(payload.leads_estimate, {"0-5": 20, "5-15": 40, "15-30": 65, "30+": 90})


def score_backlog(payload: AssessmentInput) -> int:
    if payload.backlog_days is not None:
        x = payload.backlog_days
        if x <= 1:
            return 25
        if x <= 3:
            return 45
        if x <= 7:
            return 65
        if x <= 14:
            return 82
        return 95
    return score_from_bucket(payload.backlog_bucket, {"same_day": 25, "2_3_days": 45, "4_7_days": 65, "1_2_weeks": 82, "2_plus_weeks": 95})


def score_schedule_fullness(payload: AssessmentInput) -> int:
    return score_from_bucket(payload.schedule_fullness, {"empty_slots": 25, "mixed": 55, "mostly_full": 80, "overbooked": 95})


def score_conversion(payload: AssessmentInput) -> int:
    if payload.conversion_rate is not None:
        x = payload.conversion_rate
        if x < 20:
            return 20
        if x < 30:
            return 40
        if x < 50:
            return 70
        if x < 70:
            return 90
        return 100
    bucket_score = score_from_bucket(payload.quotes_to_jobs_bucket, {"1_2": 25, "3_4": 45, "5_7": 75, "8_10": 95}, 60)
    behavior_score = score_from_bucket(payload.sales_behavior, {"lose_most": 25, "win_half": 55, "win_most": 80, "almost_everyone": 95}, 60)
    return round((bucket_score * 0.6) + (behavior_score * 0.4))


def score_response_speed(payload: AssessmentInput) -> int:
    return score_from_bucket(payload.response_speed, {"same_hour": 95, "same_day": 80, "next_day": 55, "slow": 30})


def score_utilization(payload: AssessmentInput) -> int:
    if payload.utilization_rate is not None:
        x = payload.utilization_rate
        if x < 50:
            return 25
        if x < 65:
            return 45
        if x < 80:
            return 70
        if x < 90:
            return 88
        return 95
    return score_from_bucket(payload.busy_day_description, {"fully_booked": 90, "mostly_busy": 70, "frequent_downtime": 40, "inconsistent": 30})


def score_turned_down_work(payload: AssessmentInput) -> int:
    return score_from_bucket(payload.turned_down_work, {"never": 25, "occasionally": 55, "frequently": 80, "all_the_time": 95})


def score_can_take_work_tomorrow(payload: AssessmentInput) -> int:
    return score_from_bucket(payload.can_take_work_tomorrow, {"yes_easily": 20, "somewhat": 45, "tight": 75, "no": 95})


def score_jobs_per_day(payload: AssessmentInput) -> int:
    if payload.jobs_per_crew_per_day is None:
        return 55
    x = payload.jobs_per_crew_per_day
    if x < 2:
        return 20
    if x < 4:
        return 45
    if x < 6:
        return 70
    if x < 8:
        return 85
    return 95


def score_travel_time(payload: AssessmentInput) -> int:
    if payload.travel_time_minutes is None:
        return score_from_bucket(payload.time_loss_source, {"driving": 30, "jobs_run_long": 45, "waiting": 40, "admin": 50}, 55)
    x = payload.travel_time_minutes
    if x > 45:
        return 20
    if x > 30:
        return 40
    if x > 20:
        return 60
    if x > 10:
        return 80
    return 95


def score_scheduling(payload: AssessmentInput) -> int:
    return score_from_bucket(payload.scheduling_efficiency, {"very_smooth": 95, "mostly_smooth": 75, "some_issues": 50, "chaotic": 25}, 60)


def score_admin_drag(payload: AssessmentInput) -> int:
    return score_from_bucket(payload.admin_drag, {"low": 90, "moderate": 60, "high": 30}, 60)


def score_labor_cost(payload: AssessmentInput) -> int:
    if payload.labor_cost_pct is not None:
        x = payload.labor_cost_pct
        if x < 35:
            return 90
        if x < 45:
            return 75
        if x < 55:
            return 60
        if x < 65:
            return 40
        return 20
    return 50


def score_margin(payload: AssessmentInput) -> int:
    if payload.gross_margin_pct is not None:
        x = payload.gross_margin_pct
        if x < 15:
            return 20
        if x < 25:
            return 40
        if x < 35:
            return 60
        if x < 50:
            return 80
        return 95
    return 50


def score_profitability_feeling(payload: AssessmentInput) -> int:
    return score_from_bucket(payload.profitability_feeling, {"comfortable": 90, "tight": 60, "breaking_even": 35, "struggling": 20}, 55)


def score_pricing(payload: AssessmentInput) -> int:
    return score_from_bucket(payload.pricing_adequacy, {"strong": 90, "okay": 65, "low": 30, "unclear": 50}, 50)


def score_cash_stress(payload: AssessmentInput) -> int:
    return score_from_bucket(payload.cash_stress, {"low": 90, "moderate": 60, "high": 25}, 55)


def financial_metric_number(payload: AssessmentInput, metric_key: str) -> Optional[float]:
    """Return a parsed high-confidence financial metric as a float when available."""
    metrics = payload.financial_metrics or {}
    value = metrics.get(metric_key)
    return parse_money_or_number(value)


def score_gross_margin_pct_value(value: Optional[float]) -> Optional[int]:
    if value is None:
        return None
    if value < 15:
        return 20
    if value < 25:
        return 40
    if value < 35:
        return 60
    if value < 50:
        return 80
    return 95


def score_net_margin_pct_value(value: Optional[float]) -> Optional[int]:
    if value is None:
        return None
    if value < 0:
        return 20
    if value < 3:
        return 35
    if value < 8:
        return 55
    if value < 12:
        return 70
    if value < 18:
        return 85
    return 95


def score_cash_to_debt_ratio_value(value: Optional[float]) -> Optional[int]:
    if value is None:
        return None
    if value < 0.25:
        return 25
    if value < 0.50:
        return 40
    if value < 1.00:
        return 60
    if value < 2.00:
        return 78
    return 92


def score_operating_cash_flow_value(operating_cash_flow: Optional[float], revenue: Optional[float]) -> Optional[int]:
    if operating_cash_flow is None:
        return None
    if operating_cash_flow < 0:
        return 25
    if revenue and revenue > 0:
        ocf_margin = (operating_cash_flow / revenue) * 100
        if ocf_margin < 3:
            return 45
        if ocf_margin < 8:
            return 65
        if ocf_margin < 12:
            return 80
        return 92
    return 75


def score_cash_stress_signal_value(value: Optional[str]) -> Optional[int]:
    if not value:
        return None
    return {"low": 90, "moderate": 60, "high": 25}.get(str(value).strip().lower())


def build_financial_score_details(payload: AssessmentInput) -> Dict[str, Any]:
    """
    Phase 4 financial scoring. Uses high-confidence parsed statement
    metrics first and falls back to questionnaire scoring when no parsed
    metrics are available.
    """
    metrics = payload.financial_metrics or {}
    parsed_metric_keys = {key for key in metrics.keys() if key not in {"monthly_revenue", "cash_stress_signal"}}
    components: list[dict[str, Any]] = []

    def add_component(name: str, score: Optional[int], weight: float, source: str, metric_key: Optional[str] = None, value: Any = None) -> None:
        if score is None:
            return
        components.append({
            "name": name,
            "score": clamp_score(score),
            "weight": weight,
            "source": source,
            "metric_key": metric_key,
            "value": value,
        })

    revenue = financial_metric_number(payload, "revenue")
    gross_margin_pct = financial_metric_number(payload, "gross_margin_pct")
    net_margin_pct = financial_metric_number(payload, "net_margin_pct")
    cash_to_debt_ratio = financial_metric_number(payload, "cash_to_debt_ratio")
    operating_cash_flow = financial_metric_number(payload, "net_cash_from_operations")
    cash_stress_signal = metrics.get("cash_stress_signal")

    add_component("Gross margin strength", score_gross_margin_pct_value(gross_margin_pct), 0.30, "parsed_financial_statement", "gross_margin_pct", gross_margin_pct)
    add_component("Net profitability", score_net_margin_pct_value(net_margin_pct), 0.25, "parsed_financial_statement", "net_margin_pct", net_margin_pct)
    add_component("Cash to debt strength", score_cash_to_debt_ratio_value(cash_to_debt_ratio), 0.20, "parsed_financial_statement", "cash_to_debt_ratio", cash_to_debt_ratio)
    add_component("Operating cash flow", score_operating_cash_flow_value(operating_cash_flow, revenue), 0.15, "parsed_financial_statement", "net_cash_from_operations", operating_cash_flow)
    add_component("Cash stress signal", score_cash_stress_signal_value(str(cash_stress_signal) if cash_stress_signal else None), 0.10, "parsed_financial_statement", "cash_stress_signal", cash_stress_signal)

    parsed_components = [c for c in components if c["source"] == "parsed_financial_statement"]
    if not parsed_components:
        add_component("Labor cost control", score_labor_cost(payload), 0.25, "questionnaire", "labor_cost_pct", payload.labor_cost_pct)
        add_component("Gross margin strength", score_margin(payload), 0.25, "questionnaire", "gross_margin_pct", payload.gross_margin_pct)
        add_component("Profitability feel", score_profitability_feeling(payload), 0.20, "questionnaire", "profitability_feeling", payload.profitability_feeling)
        add_component("Pricing adequacy", score_pricing(payload), 0.15, "questionnaire", "pricing_adequacy", payload.pricing_adequacy)
        add_component("Cash stress", score_cash_stress(payload), 0.15, "questionnaire", "cash_stress", payload.cash_stress)
        source_mode = "questionnaire_fallback"
    else:
        if payload.pricing_adequacy not in (None, ""):
            add_component("Pricing adequacy", score_pricing(payload), 0.05, "questionnaire_supplement", "pricing_adequacy", payload.pricing_adequacy)
        if payload.profitability_feeling not in (None, ""):
            add_component("Profitability feel", score_profitability_feeling(payload), 0.05, "questionnaire_supplement", "profitability_feeling", payload.profitability_feeling)
        source_mode = "parsed_financial_metrics"

    total_weight = sum(float(c["weight"]) for c in components)
    score = 50 if total_weight <= 0 else clamp_score(sum(float(c["score"]) * float(c["weight"]) for c in components) / total_weight)
    parsed_weight = sum(float(c["weight"]) for c in components if c["source"] == "parsed_financial_statement")
    if source_mode == "parsed_financial_metrics" and len(parsed_components) >= 3 and parsed_weight >= 0.55:
        confidence = "high"
    elif source_mode == "parsed_financial_metrics" and len(parsed_components) >= 1:
        confidence = "medium"
    else:
        confidence = "directional"

    return {
        "score": score,
        "source_mode": source_mode,
        "confidence": confidence,
        "components": components,
        "parsed_metric_keys_used": sorted(c["metric_key"] for c in components if c["source"] == "parsed_financial_statement" and c.get("metric_key")),
        "parsed_metric_keys_available": sorted(parsed_metric_keys),
        "note": "Financial score uses high-confidence parsed statement metrics when available; otherwise it falls back to questionnaire inputs.",
    }


def detect_assumptions(payload: AssessmentInput) -> list[str]:
    assumptions = []
    if payload.conversion_rate is None and not payload.quotes_to_jobs_bucket and not payload.sales_behavior:
        assumptions.append("Conversion signal is limited and partially inferred.")
    if payload.labor_cost_pct is None:
        assumptions.append("Labor cost was not provided directly.")
    if payload.gross_margin_pct is None:
        assumptions.append("Gross margin was not provided directly.")
    if payload.backlog_days is None and not payload.backlog_bucket:
        assumptions.append("Backlog was estimated from qualitative inputs.")

    # Intelligence-layer upgrades
    if payload.mode in {"guided", "hybrid"}:
        assumptions.append("Some inputs were estimated due to limited tracking systems.")
    if payload.leads_estimate and payload.leads_per_week is None:
        assumptions.append("Lead volume was estimated instead of measured directly.")
    if payload.backlog_bucket and payload.backlog_days is None:
        assumptions.append("Backlog timing was estimated from a booked-out range.")
    if payload.quotes_to_jobs_bucket and payload.conversion_rate is None:
        assumptions.append("Conversion was estimated from inquiry-to-booking buckets.")
    if payload.sales_behavior and payload.conversion_rate is None:
        assumptions.append("Conversion was inferred from sales behavior rather than a measured conversion rate.")
    if payload.busy_day_description and payload.utilization_rate is None:
        assumptions.append("Crew utilization was inferred from workload descriptions.")
    if payload.busy_day_description and payload.jobs_per_crew_per_day is None:
        assumptions.append("Capacity was estimated without exact jobs-per-day data.")
    if payload.monthly_revenue_range and payload.monthly_revenue is None:
        assumptions.append("Revenue was estimated from a range instead of exact figures.")
    if payload.time_loss_source and payload.travel_time_minutes is None:
        assumptions.append("Operational drag was inferred from the stated main time-loss source.")
    if payload.pricing_adequacy in {"unclear", None}:
        assumptions.append("Pricing strength is uncertain and may need direct validation.")
    if payload.financial_uploads_present and payload.financial_metrics:
        assumptions.append("High-confidence parsed financial statement metrics were used to strengthen the financial diagnosis.")
    elif payload.financial_uploads_present:
        assumptions.append("Financial statement files were uploaded, but structured metric extraction was limited.")

    deduped = []
    for item in assumptions:
        if item not in deduped:
            deduped.append(item)
    return deduped


def confidence_label(score: int) -> str:
    if score >= 85:
        return "High"
    if score >= 65:
        return "Medium"
    if score >= 45:
        return "Low"
    return "Exploratory"


def build_confidence_explanation(scores: Dict[str, int], assumptions: list[str], contradictions: list[str], no_constraint: bool, notes_dominant_mode: bool = False) -> str:
    reasons = []
    numeric_scores = {
        key: value
        for key, value in scores.items()
        if isinstance(value, (int, float))
    }

    if notes_dominant_mode:
        reasons.append("the report relies primarily on narrative notes because structured intake data was sparse")
    if assumptions:
        reasons.append(f"{len(assumptions)} assumption(s) were needed because some inputs were estimated or missing")
    if contradictions:
        reasons.append(f"{len(contradictions)} contradiction(s) reduced confidence in the consistency of the inputs")
    if no_constraint and not notes_dominant_mode and numeric_scores:
        reasons.append("scores were strong and tightly grouped across categories")
    elif not no_constraint and not notes_dominant_mode and len(numeric_scores) > 1:
        ordered = sorted(numeric_scores.items(), key=lambda x: x[1])
        gap = ordered[1][1] - ordered[0][1]
        reasons.append(f"the gap between the weakest and next-weakest category was {gap} point(s)")
    elif not numeric_scores:
        reasons.append("category scores were unavailable or intentionally suppressed for this assessment")
    if not reasons:
        return "Confidence is supported by fairly complete and internally consistent inputs."
    return "Confidence is influenced by " + "; ".join(reasons) + "."


def detect_contradictions(payload: AssessmentInput, scores: Dict[str, int]) -> list[str]:
    contradictions = []
    if payload.schedule_fullness == "overbooked" and payload.busy_day_description in {"frequent_downtime", "inconsistent"}:
        contradictions.append("Schedule appears overbooked, but daily capacity usage sounds inconsistent.")
    if payload.backlog_bucket in {"1_2_weeks", "2_plus_weeks"} and payload.can_take_work_tomorrow == "yes_easily":
        contradictions.append("Long backlog conflicts with ability to take work immediately.")
    if scores["demand"] > 75 and scores["capacity"] > 75 and scores["operations"] < 40:
        contradictions.append("Strong demand and capacity signals conflict with weak operational efficiency.")

    # Intelligence-layer upgrades
    if payload.turned_down_work in {"frequently", "all_the_time"} and payload.can_take_work_tomorrow == "yes_easily":
        contradictions.append("The business is turning away work despite reporting available capacity tomorrow.")
    if payload.sales_behavior in {"win_most", "almost_everyone"} and payload.conversion_rate is not None and payload.conversion_rate < 40:
        contradictions.append("Sales behavior suggests strong close rates, but the reported conversion rate is weak.")
    if payload.profitability_feeling == "comfortable" and payload.cash_stress == "high":
        contradictions.append("Profitability is described as comfortable, but cash stress is reported as high.")
    if payload.pricing_adequacy == "strong" and payload.profitability_feeling in {"breaking_even", "struggling"}:
        contradictions.append("Pricing is described as strong, yet profitability appears weak.")
    if payload.schedule_fullness == "empty_slots" and payload.turned_down_work in {"frequently", "all_the_time"}:
        contradictions.append("The schedule has open capacity, but the business also reports turning away work.")

    deduped = []
    for item in contradictions:
        if item not in deduped:
            deduped.append(item)
    return deduped


def calculate_scores(payload: AssessmentInput) -> Dict[str, Any]:
    demand = clamp_score((score_leads(payload) * 0.30) + (score_backlog(payload) * 0.35) + (score_schedule_fullness(payload) * 0.35))
    conversion = clamp_score((score_conversion(payload) * 0.80) + (score_response_speed(payload) * 0.20))
    capacity = clamp_score((score_utilization(payload) * 0.35) + (score_backlog(payload) * 0.25) + (score_turned_down_work(payload) * 0.20) + (score_can_take_work_tomorrow(payload) * 0.20))
    operations = clamp_score((score_jobs_per_day(payload) * 0.30) + (score_travel_time(payload) * 0.25) + (score_scheduling(payload) * 0.25) + (score_admin_drag(payload) * 0.20))
    financial_score_details = build_financial_score_details(payload)
    financials = financial_score_details["score"]

    scores = {"demand": demand, "conversion": conversion, "capacity": capacity, "operations": operations, "financials": financials}
    sorted_scores = sorted(scores.items(), key=lambda item: item[1])

    min_score = sorted_scores[0][1]
    max_score = sorted_scores[-1][1]
    score_spread = max_score - min_score
    notes_dominant_mode = is_notes_dominant_mode(payload)
    notes_led_mode = should_ignore_blank_structured_fields(payload)
    no_constraint = min_score >= 75 and score_spread <= 15

    if notes_led_mode or notes_dominant_mode:
        primary_constraint = "None"
        secondary_constraint = None
        no_constraint = True
    elif no_constraint:
        primary_constraint = "None"
        secondary_constraint = None
    else:
        primary_constraint = sorted_scores[0][0].capitalize()
        secondary_constraint = None
        if (sorted_scores[1][1] - sorted_scores[0][1]) <= 10:
            secondary_constraint = sorted_scores[1][0].capitalize()

    contradictions = detect_contradictions(payload, scores)
    assumptions = detect_assumptions(payload)

    completeness_fields = [
        payload.business_type, payload.location,
        payload.leads_per_week or payload.leads_estimate or payload.schedule_fullness,
        payload.conversion_rate or payload.quotes_to_jobs_bucket or payload.sales_behavior,
        payload.crews or payload.busy_day_description,
        payload.scheduling_efficiency or payload.travel_time_minutes,
        payload.monthly_revenue or payload.monthly_revenue_range or payload.profitability_feeling,
    ]
    completeness_ratio = sum(1 for x in completeness_fields if x not in (None, "", 0)) / len(completeness_fields)
    completeness_score = round(completeness_ratio * 100)

    if len(sorted_scores) > 1:
        score_gap = sorted_scores[1][1] - sorted_scores[0][1]
    else:
        score_gap = 25
    separation_score = 95 if score_gap >= 15 else 80 if score_gap >= 10 else 65 if score_gap >= 6 else 45 if score_gap >= 3 else 25
    contradiction_penalty = min(len(contradictions) * 10, 30)
    overall_confidence_score = max(20, round((completeness_score * 0.5) + (separation_score * 0.5) - contradiction_penalty))

    if notes_led_mode:
        strategic_state = "NOTES_LED"
        constraint_pattern = "The assessment is intentionally running in notes-led mode, so blank structured fields are being ignored and the narrative notes are the primary source of truth."
    elif notes_dominant_mode:
        strategic_state = "NOTES_LED"
        constraint_pattern = "The assessment is being driven primarily by the narrative notes rather than a complete structured intake, so any diagnosis should be treated as notes-led and directional."
    elif no_constraint:
        if financials >= 85:
            strategic_state = "EXIT_READY"
            constraint_pattern = "The business appears healthy, stable, and financially attractive enough to evaluate exit or acquisition opportunities."
        else:
            strategic_state = "SCALE_READY"
            constraint_pattern = "The business appears healthy across core dimensions and may be better served by scale, expansion, or strategic optimization than bottleneck fixing."
    else:
        strategic_state = "OPTIMIZATION"
        patterns = {
            "Demand": "Inconsistent lead flow limiting utilization",
            "Conversion": "Healthy inquiries but weak close effectiveness",
            "Capacity": "Demand exceeds fulfillment capacity",
            "Operations": "Scheduling and delivery inefficiency reducing output",
            "Financials": "Margin pressure limiting healthy growth",
        }
        constraint_pattern = patterns[primary_constraint]

    financial_confidence_score = 85 if financial_score_details.get("confidence") == "high" else 72 if financial_score_details.get("confidence") == "medium" else 60
    category_confidence = {"demand": 70, "conversion": 72, "capacity": 75, "operations": 68, "financials": financial_confidence_score}
    score_statuses = build_score_statuses(payload)
    display_scores = sanitize_scores_for_display(scores, score_statuses)
    score_summary_note = build_score_summary_note(payload, score_statuses)
    confidence_explanation = build_confidence_explanation(scores, assumptions, contradictions, no_constraint, notes_dominant_mode)

    return {
        "scores": display_scores,
        "raw_scores": scores,
        "score_statuses": score_statuses,
        "score_summary_note": score_summary_note,
        "category_confidence": category_confidence,
        "financial_score_details": financial_score_details,
        "primary_constraint": primary_constraint,
        "secondary_constraint": secondary_constraint,
        "constraint_pattern": constraint_pattern,
        "overall_confidence_score": overall_confidence_score,
        "overall_confidence_label": confidence_label(overall_confidence_score),
        "confidence_explanation": confidence_explanation,
        "assumptions": assumptions,
        "contradictions": contradictions,
        "no_constraint": no_constraint,
        "strategic_state": strategic_state,
    }


def build_business_type_guidance(business_type: str) -> str:
    bt = (business_type or "").strip().lower()

    keyword_groups = {
        "cleaning": ["cleaning", "janitorial", "maid", "housekeeping"],
        "lawn care": ["lawn", "landscaping", "landscape", "mowing", "yard"],
        "glass": ["glass", "mirror", "glazing", "window", "shower door"],
        "home services": ["hvac", "plumbing", "electric", "electrical", "roofing", "painting", "remodel", "garage door", "pest", "appliance", "flooring"],
        "professional services": ["consulting", "agency", "accounting", "bookkeeping", "legal", "marketing", "design", "it services", "financial advisory"],
    }

    detected = None
    for label, keywords in keyword_groups.items():
        if any(k in bt for k in keywords):
            detected = label
            break

    guidance = {
        "cleaning": [
            "Tailor actions to route density, recurring schedules, crew utilization, travel gaps, upsells, and quote follow-up.",
            "Prefer recommendations involving recurring revenue retention, zone-based scheduling, add-on services, and cleaner productivity.",
        ],
        "lawn care": [
            "Tailor actions to route density, seasonal demand swings, recurring maintenance plans, crew productivity, and weather disruption.",
            "Prefer recommendations involving route clustering, maintenance contract retention, upsells, and capacity planning during peak season.",
        ],
        "glass": [
            "Tailor actions to estimating speed, site measurement quality, install scheduling, job handoff quality, material coordination, and margin control.",
            "Prefer recommendations involving separating measurement from installation bottlenecks, tighter quote-to-install workflow, and job profitability discipline.",
        ],
        "home services": [
            "Tailor actions to dispatching, technician utilization, callback reduction, service-to-sale conversion, travel efficiency, and price discipline.",
            "Prefer recommendations involving route efficiency, faster lead response, diagnostic consistency, field productivity, and ticket-value expansion.",
        ],
        "professional services": [
            "Tailor actions to pipeline conversion, proposal follow-up, utilization, project scoping, delivery efficiency, and pricing discipline.",
            "Prefer recommendations involving tightening the sales process, reducing scope creep, improving utilization, and packaging services more profitably.",
        ],
        None: [
            "Tailor actions to the stated business type rather than giving generic service-business advice.",
            "Make the recommendations feel operationally realistic for that business model.",
        ],
    }

    lines = guidance[detected]
    detected_label = detected.title() if detected else "General Service Business"
    return "\n".join([
        f"Business-Type Context: {detected_label}",
        f"- {lines[0]}",
        f"- {lines[1]}",
    ])


SYSTEM_PROMPT = """
You are a high-level business consultant specializing in service businesses.

You are given:
- structured diagnostic scores
- identified constraint OR strategic state

Your job is to:
1. Clearly explain the situation
2. Justify it using the data
3. Translate it into business impact
4. Provide SPECIFIC, ACTIONABLE steps

CRITICAL RULES:
- If a constraint exists, focus only on that constraint
- If there is no constraint, shift to strategy: scale, optimize, or exit
- Recommendations must be practical and implementable
- Avoid vague advice like 'improve marketing'
- Tie every recommendation to a direct outcome
- Prioritize actions clearly
- Think like a consultant, not a chatbot

REPORT QUALITY RULES:
- Preserve the numbered section structure
- Preserve the separated report header
- Write in a polished, consultant-grade style
- Use stronger reasoning tied to the evidence
- Complete every section fully
- Do not end mid-sentence
""".strip()


def build_claude_user_prompt(payload: AssessmentInput, result: dict[str, Any]) -> str:
    evidence = [
        f"Business Type: {payload.business_type}",
        f"Location: {payload.location}",
        f"Mode: {payload.mode}",
        f"Requested Analysis Mode: {payload.analysis_mode}",
        f"Resolved Analysis Mode: {resolve_analysis_mode(payload)}",
        f"Notes-Led Mode: {should_ignore_blank_structured_fields(payload)}",
        "",
        "SCORES:",
        f"- Demand: {result['scores']['demand']} ({result.get('score_statuses', {}).get('demand', 'unknown')})",
        f"- Conversion: {result['scores']['conversion']} ({result.get('score_statuses', {}).get('conversion', 'unknown')})",
        f"- Capacity: {result['scores']['capacity']} ({result.get('score_statuses', {}).get('capacity', 'unknown')})",
        f"- Operations: {result['scores']['operations']} ({result.get('score_statuses', {}).get('operations', 'unknown')})",
        f"- Financials: {result['scores']['financials']} ({result.get('score_statuses', {}).get('financials', 'unknown')})",
        f"- Financial Score Details: {json.dumps(result.get('financial_score_details', {}), default=str)}",
        f"- Score Summary Note: {result.get('score_summary_note', '')}",
        "",
        f"No Constraint: {result.get('no_constraint')}",
        f"Strategic State: {result.get('strategic_state')}",
        f"Primary Constraint: {result['primary_constraint']}",
        f"Secondary Constraint: {result.get('secondary_constraint') or 'None'}",
        f"Constraint Pattern: {result.get('constraint_pattern') or 'None'}",
        f"Confidence: {result.get('overall_confidence_label', 'Unknown')} ({result.get('overall_confidence_score', 0)}/100)",
        f"Confidence Explanation: {result.get('confidence_explanation', '')}",
        "",
        "Financial Uploads:",
        f"- Financial uploads present: {payload.financial_uploads_present}",
        f"- Income Statement: {payload.income_statement_file_name or 'Not uploaded'}",
        f"- Balance Sheet: {payload.balance_sheet_file_name or 'Not uploaded'}",
        f"- Cash Flow Statement: {payload.cash_flow_statement_file_name or 'Not uploaded'}",
        f"- Financial Parse Summary: {payload.financial_parse_summary or 'No financial files parsed'}",
        f"- Financial Parse Debug: {json.dumps(payload.financial_parse_debug or {}, default=str)}",
        f"- Extracted Financial Metrics: {json.dumps(payload.financial_metrics or {}, default=str)}",
        "",
        "Narrative Notes:",
        payload.notes or "None provided",
        "",
        "Key Evidence:",
        f"- Schedule fullness: {payload.schedule_fullness}",
        f"- Backlog bucket: {payload.backlog_bucket}",
        f"- Conversion rate: {payload.conversion_rate}",
        f"- Sales behavior: {payload.sales_behavior}",
        f"- Crews: {payload.crews}",
        f"- Busy day description: {payload.busy_day_description}",
        f"- Turned down work: {payload.turned_down_work}",
        f"- Scheduling efficiency: {payload.scheduling_efficiency}",
        f"- Monthly revenue range: {payload.monthly_revenue_range}",
        f"- Profitability feeling: {payload.profitability_feeling}",
        f"- Pricing adequacy: {payload.pricing_adequacy}",
        f"- Cash stress: {payload.cash_stress}",
        "",
        "Assumptions:",
    ]
    assumptions = result.get("assumptions", [])
    evidence.extend([f"- {a}" for a in assumptions] if assumptions else ["- None"])
    evidence.extend([
        "",
        build_business_type_guidance(payload.business_type),
        "",
        "Write a client-facing consulting report in plain markdown with this exact high-level shape:",
        "Start with a divided header block:",
        "Line 1: Business name line like '<Business Type> Business Consulting Report'",
        "Line 2: Location and mode line like '<Location> | <Mode label>'",
        "",
        "Then write these numbered sections exactly in this order:",
        "1. Business Snapshot",
        "2. Diagnosis: <Primary Constraint or Strategic State>",
        "3. Why This Was Identified",
        "4. What This Means for the Business",
        "5. Risk of Inaction",
        "6. Action Plan",
        "",
        "Section rules:",
        "- Keep the numbering format exactly",
        "- Keep the tone executive and consultant-grade",
        "- Use short paragraphs and readable spacing",
        "- In section 3, use strong reasoning tied to scores and evidence",
        "- In section 6, write 3-5 action items with these sublabels:",
        "  Action X: <short title>",
        "  Time to Impact: Immediate | 30 Days | 90 Days",
        "  Effort Level: Low | Medium | High",
        "  Expected Impact: High | Medium | Foundational",
        "  What to do:",
        "  Why it matters:",
        "  Expected outcome:",
        "- Make the actions specific enough that an owner could begin execution this week",
        "- Sequence the actions from highest leverage to lowest leverage",
        "- Avoid generic advice like improve operations, improve marketing, or optimize pricing",
        "- Use concrete actions tied directly to the diagnosed constraint or strategic state",
        "- Make the action plan feel tailored to the business type, not interchangeable with other industries",
        "- If financial statement files are uploaded and parsed metrics are available, use the extracted financial metrics as stronger evidence than rough financial estimates",
        "- Use Financial Score Details to explain why the financial score is strong, weak, or directional",
        "- Do not cite unavailable or low-confidence financial metrics as evidence",
        "- Do not overstate file parsing quality; only reference extracted metrics that are present in Extracted Financial Metrics",
        "- If hard financial metrics are missing, do not guess ROI or dollar returns",
        "- When a category score is marked insufficient_data, treat it as unavailable rather than average or weak",
        "- Do not interpret blank categories as 50s or as evidence of mediocre performance",
        "- When data is qualitative or limited, still include Time to Impact and Effort Level, but keep Expected Impact directional rather than overly precise",
        "- When data is stronger, you may make Expected Impact more confident, but do not fabricate numbers",
        "- Keep sections 1 through 5 concise so section 6 has enough room for complete action plans",
        "- In section 5 or section 6, you may reference contradictions when they materially affect the recommendation",
        "- In the final lines of the report, briefly explain confidence using the provided confidence explanation",
        "- When Resolved Analysis Mode is notes_led, use the notes field as the primary source of truth for diagnosis and recommendations",
        "- When Resolved Analysis Mode is notes_led, ignore blank structured fields rather than treating them as missing, average, or weak business signals",
        "- When Resolved Analysis Mode is notes_led, do not infer a primary constraint from absent structured inputs",
        "- When Resolved Analysis Mode is notes_led, build the action plan from the narrative notes and any explicitly stated facts only",
        "",
        "Do not add new numbered sections.",
        "Do not use tables.",
        "Do not use JSON.",
        "Do not end mid-sentence.",
    ])
    return "\n".join(evidence)


def fallback_report_markdown(payload: AssessmentInput, result: Dict[str, Any], reason: str = "") -> str:
    reason_line = f"\n\n_Fallback reason: {reason}_" if reason else ""
    diagnosis = result["primary_constraint"] if not result.get("no_constraint") else result.get("strategic_state", "SCALE_READY")
    notes_led_mode = should_ignore_blank_structured_fields(payload)
    location_line = payload.location or "Location not provided"
    mode_line = f"{payload.mode.title()} Operations Model"

    return f"""
{payload.business_type} Business Consulting Report

{location_line} | {mode_line}

1. Business Snapshot

{"This assessment is intentionally running in notes-led mode, so the report is based primarily on the narrative notes and blank structured fields are being ignored." if notes_led_mode else f"This business is currently in a **{diagnosis}** state. The available inputs suggest that the most important issue should be addressed before broader growth efforts continue."}

2. Diagnosis: {diagnosis}

{"No formal structured-data constraint is being asserted because this assessment is running in notes-led mode. The diagnosis should be read as a narrative interpretation of the notes provided." if notes_led_mode else f"The model identified **{diagnosis}** as the main business issue based on the submitted operating and financial signals. The current pattern is: {result['constraint_pattern']}"}

3. Why This Was Identified

The scoring model reviewed demand, conversion, capacity, operations, and financial performance together. The lowest area or strategic-state logic determined the current diagnosis. The available evidence was strong enough to produce a directional recommendation, though confidence depends on the quality and completeness of the intake.

4. What This Means for the Business

This means the business is likely underperforming not because every part of the company is broken, but because one core issue is limiting broader performance. Until that issue is addressed, growth, profitability, or operational stability may remain constrained.

5. Risk of Inaction

If this issue is not addressed, the business may continue operating below its potential, absorb unnecessary strain, and miss the window to improve performance before the next growth cycle or seasonal shift.

6. Action Plan

Action 1: Stabilize the primary constraint  
Time to Impact: Immediate  
Effort Level: Medium  
Expected Impact: High  
What to do: Make one direct operating change within the next 7 days that targets the diagnosed bottleneck, such as tightening scheduling rules, raising price floors, or enforcing faster lead follow-up.  
Why it matters: Focused intervention on the true bottleneck creates faster results than spreading effort across multiple lower-value initiatives.  
Expected outcome: Clearer movement in throughput, close rate, booked-out time, or margin within the next two weeks.

Action 2: Convert the fix into a repeatable system  
Time to Impact: 30 Days  
Effort Level: Medium  
Expected Impact: High  
What to do: Turn the chosen fix into a documented workflow, rule, script, pricing standard, or operating routine the team can execute consistently.  
Why it matters: One-time effort creates temporary relief, but repeatable systems create sustained business improvement.  
Expected outcome: More stable execution with less owner dependence and fewer performance swings.

Action 3: Track one leading indicator every week  
Time to Impact: 30 Days  
Effort Level: Low  
Expected Impact: Foundational  
What to do: Choose the one metric most tied to the diagnosis, such as conversion rate, jobs per day, booked-out time, or revenue per job, and review it weekly.  
Why it matters: Measurement shows whether the intervention is working and helps the business correct course before issues compound.  
Expected outcome: Better decisions, faster iteration, and improved confidence in what is actually driving performance.

Assumptions: {", ".join(result["assumptions"]) if result["assumptions"] else "None"}  
Contradictions: {", ".join(result["contradictions"]) if result["contradictions"] else "None"}  
Confidence: {result["overall_confidence_label"]}  
Confidence Explanation: {result.get("confidence_explanation", "Confidence is supported by the available inputs.")}{reason_line}
""".strip()


def report_looks_incomplete(report: str) -> bool:
    if not report:
        return True

    stripped = report.strip()

    if len(stripped) < 700:
        return True

    bad_endings = (
        "You",
        "What to do:",
        "Why it matters:",
        "Expected outcome:",
        "Action 1:",
        "Action 2:",
        "Action 3:",
        "Action 4:",
        "Action 5:",
    )

    if any(stripped.endswith(x) for x in bad_endings):
        return True

    return False


def is_claude_overloaded_error(exc: Exception) -> bool:
    text = str(exc).lower()
    return "overloaded_error" in text or "overloaded" in text


def calculate_retry_delay(attempt: int, overloaded: bool = False) -> float:
    # Exponential backoff with jitter. Overload errors need longer spacing.
    base = 3.0 if overloaded else 1.75
    delay = min(30.0, base * (2 ** (attempt - 1)))
    jitter = random.uniform(0.25, 1.25)
    return delay + jitter



def build_report_markdown(payload: AssessmentInput, result: Dict[str, Any]) -> str:
    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        return fallback_report_markdown(payload, result, "Missing ANTHROPIC_API_KEY")

    model = os.getenv("ANTHROPIC_MODEL", "claude-sonnet-4-6")
    client = anthropic.Anthropic(api_key=api_key)

    prompt = build_claude_user_prompt(payload, result)
    last_error = ""

    max_attempts = 7

    for attempt in range(1, max_attempts + 1):
        try:
            print(f"=== CLAUDE CALL START attempt {attempt}/{max_attempts} ===")
            with client.messages.stream(
                model=model,
                max_tokens=3000,
                system=SYSTEM_PROMPT,
                messages=[{"role": "user", "content": prompt}],
            ) as stream:
                message = stream.get_final_message()

            parts = []
            for block in message.content:
                if getattr(block, "type", None) == "text":
                    parts.append(block.text)

            report = "\n\n".join(parts).strip()
            print(f"=== CLAUDE RESPONSE LENGTH attempt {attempt}/{max_attempts}: {len(report)} ===")
            print("=== REPORT LENGTH ===", len(report))

            if report and not report_looks_incomplete(report):
                print("=== USING CLAUDE LIVE OUTPUT ===")
                return f"# CLAUDE LIVE\n\n{report}"

            last_error = f"Incomplete Claude response on attempt {attempt}"
            print(f"=== RETRYING: {last_error} ===")
            if attempt < max_attempts:
                delay = calculate_retry_delay(attempt, overloaded=False)
                print(f"=== WAITING {delay:.1f}s BEFORE NEXT CLAUDE ATTEMPT ===")
                time.sleep(delay)

        except Exception as exc:
            overloaded = is_claude_overloaded_error(exc)
            error_type = "Claude overloaded" if overloaded else "Claude error"
            last_error = f"{error_type} on attempt {attempt}: {exc}"
            print(f"=== RETRYING: {last_error} ===")
            if attempt < max_attempts:
                delay = calculate_retry_delay(attempt, overloaded=overloaded)
                print(f"=== WAITING {delay:.1f}s BEFORE NEXT CLAUDE ATTEMPT ===")
                time.sleep(delay)

    return fallback_report_markdown(payload, result, last_error or "Claude retries exhausted")




def parse_money_or_number(value: Any) -> Optional[float]:
    """Parse accounting-style numbers safely."""
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    is_negative = text.startswith("(") and text.endswith(")")
    cleaned = re.sub(r"[^0-9.\-]", "", text)
    if cleaned in {"", "-", ".", "-."} or cleaned.count(".") > 1:
        return None
    try:
        number = float(cleaned)
        return -abs(number) if is_negative else number
    except ValueError:
        return None


def normalize_label(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip()


def rows_from_csv_or_text(path: Path) -> list[list[Any]]:
    rows: list[list[Any]] = []
    text = path.read_text(encoding="utf-8", errors="ignore")
    if "<table" in text.lower() or "<tr" in text.lower():
        tr_blocks = re.findall(r"<tr[^>]*>(.*?)</tr>", text, flags=re.IGNORECASE | re.DOTALL)
        for tr in tr_blocks:
            cells = re.findall(r"<t[dh][^>]*>(.*?)</t[dh]>", tr, flags=re.IGNORECASE | re.DOTALL)
            cleaned_cells = [re.sub(r"<[^>]+>", "", c).strip() for c in cells]
            if any(cleaned_cells):
                rows.append(cleaned_cells)
        return rows
    sample = text[:2048]
    delimiter = "\t" if "\t" in sample and sample.count("\t") > sample.count(",") else ","
    reader = csv.reader(text.splitlines(), delimiter=delimiter)
    for row in reader:
        if any(str(cell).strip() for cell in row):
            rows.append(row)
    return rows


def rows_from_xlsx(path: Path) -> list[list[Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []
    wb = load_workbook(path, data_only=True, read_only=True)
    rows: list[list[Any]] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            values = list(row)
            if any(v not in (None, "") for v in values):
                rows.append(values)
    return rows


def extract_pdf_text(path: Path) -> str:
    text = ""
    try:
        from pypdf import PdfReader
        reader = PdfReader(str(path))
        for page in reader.pages:
            text += "\n" + (page.extract_text() or "")
        if text.strip():
            return text
    except Exception:
        pass
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(str(path))
        for page in reader.pages:
            text += "\n" + (page.extract_text() or "")
    except Exception:
        return ""
    return text


def split_pdf_label_and_values(line: str) -> Optional[list[Any]]:
    cleaned = re.sub(r"\s+", " ", line.strip())
    if not cleaned:
        return None
    money_pattern = r"\(?\$?\s*-?\d[\d,]*(?:\.\d+)?\)?"
    matches = list(re.finditer(money_pattern, cleaned))
    if not matches:
        return [cleaned]
    first_number = matches[0]
    label = cleaned[: first_number.start()].strip(" -:\t")
    if not label:
        return None
    values = [m.group(0).strip() for m in matches]
    trailing = cleaned[matches[-1].end():].strip(" -:\t")
    row: list[Any] = [label, *values]
    if trailing:
        row.append(trailing)
    return row


def rows_from_pdf_text_lines(text: str) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for raw_line in text.splitlines():
        parsed = split_pdf_label_and_values(raw_line.strip())
        if parsed and len(parsed) > 1:
            rows.append(parsed)
    return rows


def rows_from_pdf_column_text(text: str) -> list[list[Any]]:
    lines = [re.sub(r"\s+", " ", line.strip()) for line in text.splitlines() if line.strip()]
    rows: list[list[Any]] = []
    money_pattern = re.compile(r"^\(?\$?\s*-?\d[\d,]*(?:\.\d+)?\)?$")
    i = 0
    while i < len(lines):
        if not metric_key_from_label(lines[i]):
            i += 1
            continue
        values: list[str] = []
        j = i + 1
        while j < len(lines) and len(values) < 5:
            if money_pattern.match(lines[j]):
                values.append(lines[j]); j += 1; continue
            if metric_key_from_label(lines[j]):
                break
            if values:
                break
            j += 1
        if values:
            rows.append([lines[i], *values]); i = j
        else:
            i += 1
    return rows


def rows_from_pdf(path: Path) -> list[list[Any]]:
    text = extract_pdf_text(path)
    if not text.strip():
        return []
    rows = rows_from_pdf_text_lines(text)
    recognized_rows = [row for row in rows if row and metric_key_from_label(str(row[0]))]
    if len(recognized_rows) < 3:
        column_rows = rows_from_pdf_column_text(text)
        if len(column_rows) > len(recognized_rows):
            rows = column_rows
    return rows


def rows_from_financial_file(path: Path) -> list[list[Any]]:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".xls"}:
        return rows_from_csv_or_text(path)
    if suffix == ".xlsx":
        return rows_from_xlsx(path)
    if suffix == ".pdf":
        return rows_from_pdf(path)
    return []


def looks_like_year_number(number: float) -> bool:
    return 1900 <= abs(number) <= 2100 and float(number).is_integer()


def cell_year(value: Any) -> Optional[int]:
    text = str(value or "").strip()
    match = re.search(r"\b(19\d{2}|20\d{2}|21\d{2})\b", text)
    return int(match.group(1)) if match else None


def cell_has_financial_number(value: Any) -> bool:
    number = parse_money_or_number(value)
    return number is not None and not looks_like_year_number(number)


FINANCIAL_LABEL_ALIASES: dict[str, list[str]] = {
    "revenue": ["total revenue", "service revenue", "revenue", "sales", "net sales"],
    "cogs": ["cost of goods sold", "cogs", "cost of sales", "cost of revenue", "total cost of revenue", "direct costs"],
    "gross_profit": ["gross profit"],
    "operating_expenses": ["total operating expenses", "operating expenses", "opex", "total expenses"],
    "operating_income": ["operating income", "income from operations", "ebit"],
    "interest_expense": ["interest expense"],
    "net_income": ["net income", "net profit"],
    "cash": ["cash", "cash and cash equivalents", "cash equivalents", "ending cash"],
    "accounts_receivable": ["accounts receivable", "trade receivables", "ar", "a r"],
    "inventory": ["inventory"],
    "total_assets": ["total assets"],
    "accounts_payable": ["accounts payable", "trade payables", "ap", "a p"],
    "debt": ["total debt", "current debt", "long term debt", "long-term debt", "debt", "loans", "notes payable"],
    "owner_equity": ["owner equity", "owners equity", "total equity", "equity", "retained earnings"],
    "net_cash_from_operations": ["net cash from operations", "net cash provided by operating activities", "operating cash flow", "cash from operations"],
    "capital_expenditures": ["capital expenditures", "capex", "purchases of property and equipment"],
    "debt_repayment": ["debt repayment", "loan repayment", "principal repayment", "repayments of debt"],
    "net_change_in_cash": ["net change in cash", "change in cash", "net increase in cash", "net decrease in cash"],
}

STATEMENT_METRIC_RULES: dict[str, dict[str, dict[str, Any]]] = {
    "income_statement": {
        "revenue": {"aliases": ["total revenue", "net sales", "total sales", "revenue", "service revenue"], "exclude": ["materials revenue", "product revenue"], "prefer_total": True},
        "cogs": {"aliases": ["cost of goods sold", "total cost of revenue", "cost of revenue", "cost of sales", "cogs", "direct costs"], "exclude": ["direct labor", "materials and supplies", "subcontractors"], "prefer_total": True},
        "gross_profit": {"aliases": ["gross profit"], "exclude": [], "prefer_total": False},
        "operating_expenses": {"aliases": ["total operating expenses", "operating expenses", "opex"], "exclude": ["office payroll", "rent", "marketing", "insurance", "software"], "prefer_total": True},
        "operating_income": {"aliases": ["operating income", "income from operations", "ebit"], "exclude": [], "prefer_total": False},
        "interest_expense": {"aliases": ["interest expense"], "exclude": [], "prefer_total": False},
        "net_income": {"aliases": ["net income", "net profit"], "exclude": [], "prefer_total": False},
    },
    "balance_sheet": {
        "cash": {"aliases": ["cash", "cash and cash equivalents", "cash equivalents"], "exclude": [], "prefer_total": False},
        "accounts_receivable": {"aliases": ["accounts receivable", "trade receivables", "a r", "ar"], "exclude": ["change in accounts receivable"], "prefer_total": False},
        "inventory": {"aliases": ["inventory"], "exclude": ["change in inventory"], "prefer_total": False},
        "total_assets": {"aliases": ["total assets"], "exclude": ["total current assets"], "prefer_total": True},
        "accounts_payable": {"aliases": ["accounts payable", "trade payables", "a p", "ap"], "exclude": ["change in accounts payable"], "prefer_total": False},
        "debt": {"aliases": ["total debt", "current debt", "long term debt", "long-term debt", "notes payable", "loans"], "exclude": ["debt repayment", "debt issued"], "prefer_total": True, "aggregate_components": True},
        "owner_equity": {"aliases": ["owner equity", "owners equity", "total equity", "equity", "retained earnings"], "exclude": [], "prefer_total": False},
    },
    "cash_flow_statement": {
        "net_income": {"aliases": ["net income"], "exclude": [], "prefer_total": False},
        "net_cash_from_operations": {"aliases": ["net cash from operations", "net cash provided by operating activities", "operating cash flow", "cash from operations"], "exclude": [], "prefer_total": True},
        "capital_expenditures": {"aliases": ["capital expenditures", "capex", "purchases of property and equipment"], "exclude": [], "prefer_total": False},
        "debt_repayment": {"aliases": ["debt repayment", "loan repayment", "principal repayment", "repayments of debt"], "exclude": ["debt issued"], "prefer_total": False},
        "net_change_in_cash": {"aliases": ["net change in cash", "change in cash", "net increase in cash", "net decrease in cash"], "exclude": [], "prefer_total": False},
        "cash": {"aliases": ["ending cash", "cash at end of period"], "exclude": ["net change"], "prefer_total": False},
    },
}


def metric_key_from_label(label: str) -> Optional[str]:
    normalized = normalize_label(label)
    if not normalized:
        return None
    best_key: Optional[str] = None
    best_score = -1
    for key, aliases in FINANCIAL_LABEL_ALIASES.items():
        for alias in aliases:
            normalized_alias = normalize_label(alias)
            score = -1
            if normalized == normalized_alias:
                score = 1000 + len(normalized_alias)
            elif normalized.endswith(normalized_alias):
                score = 700 + len(normalized_alias)
            elif normalized_alias in normalized:
                score = 400 + len(normalized_alias)
            if score > best_score:
                best_key, best_score = key, score
    return best_key


def detect_table_structure(rows: list[list[Any]]) -> dict[str, Any]:
    best_header_index: Optional[int] = None
    best_year_columns: dict[int, int] = {}
    best_score = -1
    for idx, row in enumerate(rows[:15]):
        year_columns = {col_idx: cell_year(cell) for col_idx, cell in enumerate(row) if cell_year(cell) is not None}
        text_cells = sum(1 for cell in row if normalize_label(cell) and not cell_year(cell))
        score = len(year_columns) * 10 + text_cells
        if year_columns and score > best_score:
            best_header_index, best_year_columns, best_score = idx, year_columns, score
    if best_year_columns:
        latest_year_col = max(best_year_columns, key=lambda c: best_year_columns[c])
        first_value_col = min(best_year_columns)
        return {"header_row": best_header_index, "year_columns": best_year_columns, "latest_value_column": latest_year_col, "latest_year": best_year_columns[latest_year_col], "label_columns": list(range(first_value_col)) or [0], "single_period": False}
    label_width_counts: dict[int, int] = {}
    for row in rows:
        numeric_cols = [i for i, cell in enumerate(row) if cell_has_financial_number(cell)]
        if numeric_cols and min(numeric_cols) > 0:
            label_width_counts[min(numeric_cols)] = label_width_counts.get(min(numeric_cols), 0) + 1
    label_width = max(label_width_counts, key=label_width_counts.get) if label_width_counts else 1
    return {"header_row": None, "year_columns": {}, "latest_value_column": None, "latest_year": None, "label_columns": list(range(max(1, label_width))), "single_period": True}


def row_label_parts(row: list[Any], structure: dict[str, Any]) -> tuple[str, str, str]:
    labels = [str(row[i]).strip() for i in (structure.get("label_columns") or [0]) if i < len(row) and str(row[i] or "").strip()]
    category = labels[0] if len(labels) > 1 else ""
    line_item = labels[-1] if labels else ""
    return category, line_item, " ".join(labels).strip()


def value_for_row(row: list[Any], structure: dict[str, Any]) -> tuple[Optional[float], Optional[int], Optional[str], float]:
    latest_col = structure.get("latest_value_column")
    latest_year = structure.get("latest_year")
    if latest_col is not None and latest_col < len(row):
        value = parse_money_or_number(row[latest_col])
        if value is not None and not looks_like_year_number(value):
            return value, latest_col, str(latest_year) if latest_year else None, 0.98
    numeric_cells: list[tuple[int, float]] = []
    max_label_col = max(structure.get("label_columns") or [0])
    for idx, cell in enumerate(row):
        if idx <= max_label_col:
            continue
        number = parse_money_or_number(cell)
        if number is not None and not looks_like_year_number(number):
            numeric_cells.append((idx, number))
    if numeric_cells:
        col_idx, value = numeric_cells[-1]
        return value, col_idx, None, 0.72 if structure.get("single_period") else 0.62
    return None, None, None, 0.0


def alias_match_score(label: str, aliases: list[str], exclusions: list[str]) -> tuple[float, Optional[str]]:
    normalized = normalize_label(label)
    if not normalized:
        return 0.0, None
    for exclusion in exclusions:
        normalized_exclusion = normalize_label(exclusion)
        if normalized_exclusion and normalized_exclusion in normalized:
            return 0.0, None
    best_score, best_alias = 0.0, None
    for alias in aliases:
        normalized_alias = normalize_label(alias)
        if normalized == normalized_alias:
            score = 1.0
        elif normalized.endswith(normalized_alias):
            score = 0.92
        elif normalized_alias in normalized:
            score = 0.80
        else:
            score = 0.0
        if score > best_score:
            best_score, best_alias = score, alias
    return best_score, best_alias


def confidence_for_candidate(alias_score: float, value_confidence: float, statement_type: str, metric_key: str, category: str, line_item: str, structure: dict[str, Any], rule: dict[str, Any]) -> float:
    confidence = alias_score * 0.65 + value_confidence * 0.25
    if structure.get("latest_year") is not None:
        confidence += 0.07
    if statement_type in STATEMENT_METRIC_RULES:
        confidence += 0.05
    normalized_line = normalize_label(line_item)
    normalized_category = normalize_label(category)
    if rule.get("prefer_total") and "total" in normalized_line:
        confidence += 0.06
    if metric_key == "debt" and any(x in normalized_line for x in ["debt", "loan", "notes payable"]):
        confidence += 0.04
    if metric_key == "total_assets" and normalized_line == "total assets":
        confidence += 0.08
    if metric_key == "revenue" and normalized_category == "revenue" and "total" in normalized_line:
        confidence += 0.06
    return round(min(confidence, 0.99), 3)


def extract_statement_metrics_from_rows(rows: list[list[Any]], statement_type: str, file_name: str = "") -> dict[str, Any]:
    structure = detect_table_structure(rows)
    rules = STATEMENT_METRIC_RULES.get(statement_type, {})
    candidates_by_metric: dict[str, list[dict[str, Any]]] = {key: [] for key in rules.keys()}
    debug: dict[str, Any] = {}
    row_start = (structure.get("header_row") + 1) if structure.get("header_row") is not None else 0
    for row_index, row in enumerate(rows[row_start:], start=row_start + 1):
        if not row:
            continue
        category, line_item, combined_label = row_label_parts(row, structure)
        label_for_matching = line_item or combined_label
        if not normalize_label(label_for_matching):
            continue
        value, value_col, column_used, value_confidence = value_for_row(row, structure)
        if value is None:
            continue
        for metric_key, rule in rules.items():
            alias_score, alias_used = alias_match_score(label_for_matching, rule.get("aliases", []), rule.get("exclude", []))
            if alias_score <= 0:
                alias_score, alias_used = alias_match_score(combined_label, rule.get("aliases", []), rule.get("exclude", []))
            if alias_score <= 0:
                continue
            confidence = confidence_for_candidate(alias_score, value_confidence, statement_type, metric_key, category, line_item, structure, rule)
            candidates_by_metric.setdefault(metric_key, []).append({
                "metric": metric_key, "value": value, "source_file": file_name, "row": row_index,
                "label_used": label_for_matching, "category": category, "alias_used": alias_used,
                "value_selected": value, "value_column_index": value_col,
                "column_used": column_used or ("single_period_or_rightmost_value" if structure.get("single_period") else None),
                "confidence": confidence,
            })
    metrics: dict[str, float] = {}
    unavailable: dict[str, Any] = {}
    threshold = 0.78
    for metric_key, candidates in candidates_by_metric.items():
        if not candidates:
            unavailable[metric_key] = {"reason": "not_found", "confidence": 0.0, "source_file": file_name}
            continue
        candidates.sort(key=lambda c: (c["confidence"], "total" in normalize_label(c.get("label_used")), abs(float(c.get("value") or 0)) if rules.get(metric_key, {}).get("prefer_total") else 0), reverse=True)
        best = candidates[0]
        if statement_type == "balance_sheet" and metric_key == "debt":
            total_debt_candidates = [c for c in candidates if "total debt" in normalize_label(c.get("label_used"))]
            component_candidates = [c for c in candidates if any(x in normalize_label(c.get("label_used")) for x in ["current debt", "long term debt", "long term debt", "long term debt"]) and c["confidence"] >= threshold]
            if total_debt_candidates and total_debt_candidates[0]["confidence"] >= threshold:
                best = total_debt_candidates[0]
            elif len(component_candidates) >= 2:
                component_sum = sum(float(c["value"]) for c in component_candidates)
                best = {"metric": "debt", "value": component_sum, "source_file": file_name, "row": [c["row"] for c in component_candidates], "label_used": " + ".join(c["label_used"] for c in component_candidates), "category": "Current Liabilities / Long-Term Liabilities", "alias_used": "debt components", "value_selected": component_sum, "value_column_index": [c["value_column_index"] for c in component_candidates], "column_used": component_candidates[0].get("column_used"), "confidence": round(min(min(c["confidence"] for c in component_candidates) + 0.03, 0.96), 3), "aggregation": "sum_of_current_debt_and_long_term_debt"}
        if best["confidence"] >= threshold:
            metrics[metric_key] = float(best["value"])
            debug[metric_key] = best
        else:
            unavailable[metric_key] = {"reason": "low_confidence", "source_file": file_name, "row": best.get("row"), "label_used": best.get("label_used"), "value_selected": best.get("value_selected"), "confidence": best.get("confidence", 0.0)}
    return {"metrics": metrics, "debug": debug, "unavailable_metrics": unavailable, "structure": structure}


def extract_metrics_from_rows(rows: list[list[Any]], statement_type: str = "unknown", file_name: str = "") -> dict[str, float]:
    return extract_statement_metrics_from_rows(rows, statement_type, file_name).get("metrics", {})


def derive_financial_metrics(metrics: dict[str, Any]) -> dict[str, Any]:
    derived = dict(metrics)
    revenue = parse_money_or_number(derived.get("revenue"))
    cogs = parse_money_or_number(derived.get("cogs"))
    gross_profit = parse_money_or_number(derived.get("gross_profit"))
    net_income = parse_money_or_number(derived.get("net_income"))
    cash = parse_money_or_number(derived.get("cash"))
    debt = parse_money_or_number(derived.get("debt"))
    net_cash_from_operations = parse_money_or_number(derived.get("net_cash_from_operations"))
    if gross_profit is None and revenue is not None and cogs is not None:
        gross_profit = revenue - cogs
        derived["gross_profit"] = gross_profit
    if revenue and gross_profit is not None:
        derived["gross_margin_pct"] = round((gross_profit / revenue) * 100, 1)
    if revenue:
        derived["monthly_revenue"] = round(revenue / 12, 2)
    if revenue and net_income is not None:
        derived["net_margin_pct"] = round((net_income / revenue) * 100, 1)
    if cash is not None and debt is not None:
        derived["cash_to_debt_ratio"] = round(cash / debt, 2) if debt else None
    if net_cash_from_operations is not None and net_cash_from_operations < 0:
        derived["cash_stress_signal"] = "high"
    elif cash is not None and debt is not None and debt > cash * 2:
        derived["cash_stress_signal"] = "moderate"
    elif cash is not None:
        derived["cash_stress_signal"] = "low"
    return derived


def parse_financial_uploads(uploads: dict[str, Optional[dict[str, str]]]) -> dict[str, Any]:
    parsed: dict[str, Any] = {"statements": {}, "combined_metrics": {}, "financial_parse_debug": {}, "unavailable_metrics": {}, "parse_errors": []}
    for statement_type, metadata in uploads.items():
        if not metadata:
            continue
        path = Path(metadata["file_path"])
        try:
            rows = rows_from_financial_file(path)
            extraction = extract_statement_metrics_from_rows(rows, statement_type, metadata.get("file_name", path.name))
            metrics = extraction.get("metrics", {})
            debug = extraction.get("debug", {})
            unavailable = extraction.get("unavailable_metrics", {})
            parsed["statements"][statement_type] = {"file_name": metadata["file_name"], "row_count": len(rows), "metrics": metrics, "financial_parse_debug": debug, "unavailable_metrics": unavailable, "table_structure": extraction.get("structure", {})}
            parsed["combined_metrics"].update(metrics)
            parsed["financial_parse_debug"].update({f"{statement_type}.{k}": v for k, v in debug.items()})
            parsed["unavailable_metrics"].update({f"{statement_type}.{k}": v for k, v in unavailable.items()})
        except Exception as exc:
            parsed["parse_errors"].append(f"{statement_type}: {exc}")
    parsed["combined_metrics"] = derive_financial_metrics(parsed["combined_metrics"])
    return parsed


def apply_parsed_financial_metrics(payload_dict: dict[str, Any], parsed_financials: dict[str, Any]) -> dict[str, Any]:
    metrics = parsed_financials.get("combined_metrics", {}) or {}
    if payload_dict.get("monthly_revenue") in (None, "", 0) and metrics.get("monthly_revenue") is not None:
        payload_dict["monthly_revenue"] = metrics["monthly_revenue"]
    if payload_dict.get("gross_margin_pct") in (None, "", 0) and metrics.get("gross_margin_pct") is not None:
        payload_dict["gross_margin_pct"] = metrics["gross_margin_pct"]
    if payload_dict.get("cash_stress") in (None, "") and metrics.get("cash_stress_signal") is not None:
        payload_dict["cash_stress"] = metrics["cash_stress_signal"]
    payload_dict["financial_metrics"] = metrics
    payload_dict["financial_parse_debug"] = parsed_financials.get("financial_parse_debug", {})
    extracted = sorted(metrics.keys())
    unavailable_count = len(parsed_financials.get("unavailable_metrics", {}) or {})
    errors = parsed_financials.get("parse_errors", [])
    payload_dict["financial_parse_summary"] = (f"Extracted {len(extracted)} high-confidence financial metric(s): {', '.join(extracted[:12])}" if extracted else "Financial files were uploaded, but no high-confidence structured metrics were extracted.")
    if unavailable_count:
        payload_dict["financial_parse_summary"] += f" {unavailable_count} metric(s) were marked unavailable because they were missing or low confidence."
    if errors:
        payload_dict["financial_parse_summary"] += f" Parse warnings: {'; '.join(errors)}"
    return payload_dict

UPLOAD_DIR = Path(os.getenv("FINANCIAL_UPLOAD_DIR", "uploaded_financials"))


def allowed_financial_upload(filename: str) -> bool:
    return Path(filename).suffix.lower() in {".csv", ".xlsx", ".xls", ".pdf"}


async def save_financial_upload(assessment_upload_id: str, statement_type: str, upload: Optional[UploadFile]) -> Optional[dict[str, str]]:
    if upload is None or not upload.filename:
        return None

    if not allowed_financial_upload(upload.filename):
        raise HTTPException(status_code=400, detail=f"Unsupported file type for {statement_type}. Upload CSV, Excel, or PDF files only.")

    upload_folder = UPLOAD_DIR / assessment_upload_id
    upload_folder.mkdir(parents=True, exist_ok=True)

    safe_name = Path(upload.filename).name
    saved_path = upload_folder / f"{statement_type}_{safe_name}"
    saved_path.write_bytes(await upload.read())

    return {"file_name": safe_name, "file_path": str(saved_path)}


def apply_financial_upload_metadata(payload_dict: dict[str, Any], uploads: dict[str, Optional[dict[str, str]]]) -> dict[str, Any]:
    mapping = {
        "income_statement": ("income_statement_file_name", "income_statement_file_path"),
        "balance_sheet": ("balance_sheet_file_name", "balance_sheet_file_path"),
        "cash_flow_statement": ("cash_flow_statement_file_name", "cash_flow_statement_file_path"),
    }

    any_uploads = False
    for upload_key, metadata in uploads.items():
        if metadata:
            name_field, path_field = mapping[upload_key]
            payload_dict[name_field] = metadata["file_name"]
            payload_dict[path_field] = metadata["file_path"]
            any_uploads = True

    payload_dict["financial_uploads_present"] = any_uploads or bool(payload_dict.get("financial_uploads_present"))
    return payload_dict

@app.get("/health")
def health() -> dict[str, Any]:
    return {
        "status": "ok",
        "app": "AI Business Diagnostic API",
        "model_version": "v2",
        "storage": "sqlite_enabled",
        "claude_configured": bool(os.getenv("ANTHROPIC_API_KEY")),
        "anthropic_model": os.getenv("ANTHROPIC_MODEL", "claude-sonnet-4-6"),
        "prompt_version": os.getenv("PROMPT_VERSION", "report_v2_strategic"),
        "analysis_modes_supported": ["auto", "structured", "hybrid", "notes_led"],
        "financial_upload_phase": "phase_4_metrics_feed_scoring_and_reports",
    }


@app.post("/api/v1/assessments/score")
def score_assessment(payload: AssessmentInput) -> dict[str, Any]:
    result = calculate_scores(payload)
    return {"assessment_id": f"preview-{uuid4()}", **result, "report_markdown": ""}


@app.post("/api/v1/assessments/report")
def create_report_route(payload: AssessmentInput, db: Session = Depends(get_db)) -> dict[str, Any]:
    result = calculate_scores(payload)
    report_markdown = build_report_markdown(payload, result)
    assessment = create_assessment(
        db,
        payload_dict=payload.model_dump(),
        result=result,
        report_markdown=report_markdown,
        scoring_model_version="v2",
        prompt_version=os.getenv("PROMPT_VERSION", "report_v2_strategic"),
    )
    latest_report = assessment.reports[-1]
    return {"assessment_id": assessment.id, **result, "report_markdown": latest_report.report_markdown}


@app.post("/api/v1/assessments/report-with-files")
async def create_report_with_files_route(
    payload_json: str = Form(...),
    income_statement: Optional[UploadFile] = File(default=None),
    balance_sheet: Optional[UploadFile] = File(default=None),
    cash_flow_statement: Optional[UploadFile] = File(default=None),
    db: Session = Depends(get_db),
) -> dict[str, Any]:
    try:
        payload_dict = json.loads(payload_json)
        upload_id = str(uuid4())

        uploads = {
            "income_statement": await save_financial_upload(upload_id, "income_statement", income_statement),
            "balance_sheet": await save_financial_upload(upload_id, "balance_sheet", balance_sheet),
            "cash_flow_statement": await save_financial_upload(upload_id, "cash_flow_statement", cash_flow_statement),
        }

        payload_dict = apply_financial_upload_metadata(payload_dict, uploads)
        parsed_financials = parse_financial_uploads(uploads)
        payload_dict = apply_parsed_financial_metrics(payload_dict, parsed_financials)
        payload = AssessmentInput(**payload_dict)

        result = calculate_scores(payload)
        result["financial_uploads"] = uploads
        result["financial_uploads_present"] = payload.financial_uploads_present
        result["financial_metrics"] = payload.financial_metrics or {}
        result["financial_parse_summary"] = payload.financial_parse_summary
        result["financial_parse_debug"] = payload.financial_parse_debug or {}
        result["parsed_financials"] = parsed_financials

        report_markdown = build_report_markdown(payload, result)
        assessment = create_assessment(
            db,
            payload_dict=payload_dict,
            result=result,
            report_markdown=report_markdown,
            scoring_model_version="v2",
            prompt_version=os.getenv("PROMPT_VERSION", "report_v2_strategic"),
        )
        latest_report = assessment.reports[-1]
        return {"assessment_id": assessment.id, **result, "report_markdown": latest_report.report_markdown}
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to process report with files: {exc}")


@app.get("/api/v1/assessments/{assessment_id}")
def get_assessment_route(assessment_id: str, db: Session = Depends(get_db)) -> dict[str, Any]:
    assessment = get_assessment_by_id(db, assessment_id)
    if not assessment:
        raise HTTPException(status_code=404, detail="Not Found")
    latest_report = assessment.reports[-1] if assessment.reports else None

    no_constraint = assessment.primary_constraint == "None"
    strategic_state = "OPTIMIZATION"
    if no_constraint:
        if assessment.constraint_pattern and "notes-led mode" in assessment.constraint_pattern.lower():
            strategic_state = "NOTES_LED"
        elif assessment.constraint_pattern and "notes" in assessment.constraint_pattern.lower():
            strategic_state = "NOTES_LED"
        elif assessment.constraint_pattern and "exit" in assessment.constraint_pattern.lower():
            strategic_state = "EXIT_READY"
        else:
            strategic_state = "SCALE_READY"

    raw_scores = json.loads(assessment.scores_json)
    assumptions = json.loads(assessment.assumptions_json)
    contradictions = json.loads(assessment.contradictions_json)
    notes_dominant_mode = "notes" in (assessment.constraint_pattern or "").lower() or strategic_state == "NOTES_LED"
    score_statuses = {k: ("insufficient_data" if notes_dominant_mode else "estimated") for k in raw_scores.keys()}
    display_scores = {k: (None if v == "insufficient_data" else raw_scores[k]) for k, v in score_statuses.items()}
    score_summary_note = "Limited structured input was provided, so category scores are being treated as unavailable while the report relies primarily on the notes field." if notes_dominant_mode else "Stored category scores may include estimated values from partial structured inputs."
    confidence_explanation = build_confidence_explanation(raw_scores, assumptions, contradictions, no_constraint, notes_dominant_mode)

    return {
        "assessment_id": assessment.id,
        "scores": display_scores,
        "raw_scores": raw_scores,
        "score_statuses": score_statuses,
        "score_summary_note": score_summary_note,
        "category_confidence": json.loads(assessment.category_confidence_json),
        "primary_constraint": assessment.primary_constraint,
        "secondary_constraint": assessment.secondary_constraint,
        "constraint_pattern": assessment.constraint_pattern,
        "overall_confidence_score": assessment.overall_confidence_score,
        "overall_confidence_label": assessment.overall_confidence_label,
        "confidence_explanation": confidence_explanation,
        "assumptions": assumptions,
        "contradictions": contradictions,
        "no_constraint": no_constraint,
        "strategic_state": strategic_state,
        "report_markdown": latest_report.report_markdown if latest_report else "",
    }


@app.get("/api/v1/assessments")
def list_assessments_route(limit: int = Query(default=20, ge=1, le=100), db: Session = Depends(get_db)) -> list[dict[str, Any]]:
    rows = list_recent_assessments(db, limit=limit)
    return [
        {
            "assessment_id": row.id,
            "business_type": row.business_type,
            "location": row.location,
            "primary_constraint": row.primary_constraint,
            "secondary_constraint": row.secondary_constraint,
            "overall_confidence_label": row.overall_confidence_label,
            "overall_confidence_score": row.overall_confidence_score,
            "created_at": row.created_at.isoformat() if row.created_at else None,
        }
        for row in rows
    ]
